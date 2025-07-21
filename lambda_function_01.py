# ======================================
# 標準ライブラリ
# ======================================
import json
import logging
from datetime import datetime, timedelta
from io import BytesIO

# ======================================
# サードパーティ製ライブラリ
# ======================================
import boto3
from openpyxl import Workbook, load_workbook

# ======================================
# ログ設定
# ======================================
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# ======================================
# AWSクライアントの初期化
# ======================================
connect_client = boto3.client('connect')
s3_client = boto3.client('s3')

# ======================================
# 定数設定
# ======================================
BUCKET_NAME = 'amazon-connect-list'         # 対象のS3バケット名
INSTANCE_ID = '**********'                  # Amazon ConnectのインスタンスID
OBJECT_KEY = 'シフト表.xlsx'                # シフト表が保存されているExcelファイルのS3キー
CALL_LOG_FOLDER = 'connect-call-log/'       # 通話ログの保存先フォルダ


# ======================================
# Excelの日付セルを日付型に変換する関数
# ======================================
def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


# ======================================
# Lambda関数のメイン処理
# ======================================
def lambda_handler(event, context):
    try:
        logger.info(f"受信イベント: {json.dumps(event)}")

        # イベントから現在の連絡先インデックスを取得（初回はNone）
        contact_index = event.get("next_index", None)

        # 本日のUTC日付を取得
        alarm_date = datetime.utcnow().date()
        logger.info(f"アラーム発生日: {alarm_date}")

        # ======================================
        # S3からシフトExcelファイルを取得して読み込み
        # ======================================
        response = s3_client.get_object(Bucket=BUCKET_NAME, Key=OBJECT_KEY)
        excel_data = response['Body'].read()
        wb = load_workbook(filename=BytesIO(excel_data), data_only=True)
        sheet = wb['シフト']

        # ======================================
        # Excelから発信候補者を抽出
        # ======================================
        all_candidates = []
        for row in sheet.iter_rows(min_row=2):  # ヘッダーを除いて2行目以降を処理
            raw_phone = str(row[8].value).strip() if row[8].value else None
            name = str(row[0].value).strip() if row[0].value else "不明"
            row_date = parse_excel_date(row[3].value)
            if raw_phone:
                digits_only = ''.join(filter(str.isdigit, raw_phone))  # 数字のみ抽出
                if digits_only.startswith('0') and len(digits_only) >= 10:
                    # 日本の電話番号をE.164形式に変換（例: 09012345678 → +819012345678）
                    e164_phone = '+81' + digits_only[1:]
                    all_candidates.append((name, e164_phone, row_date))

        logger.info(f"候補者総数: {len(all_candidates)}")

        # ======================================
        # 発信対象者を決定
        # ======================================
        if contact_index is None:
            # 初回発信：本日の日付に一致する行のみ対象
            candidates = [(n, p) for n, p, d in all_candidates if d == alarm_date]
            contact_index = 0
        else:
            # 2回目以降：日付に関係なく全て対象
            candidates = [(n, p) for n, p, _ in all_candidates]

        logger.info(f"発信候補数: {len(candidates)} (index={contact_index})")

        if contact_index >= len(candidates):
            # 発信対象者をすべて処理済みの場合
            logger.info("全員に発信済。終了します。")
            return {"statusCode": 200, "body": "全員に発信済"}

        # 現在のインデックスの人に発信
        name, phone_number = candidates[contact_index]
        now_str = datetime.now().isoformat()

        # Amazon Connectの発信時に送るカスタム属性
        custom_attributes = {'Message': 'アラーム発生'}
        contact_flow_id = '**********'  # ConnectのコンタクトフローID

        # ======================================
        # 通話ログExcelを準備
        # ======================================
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_key = f'{CALL_LOG_FOLDER}result_log_{timestamp}.xlsx'

        wb_log = Workbook()
        ws_log = wb_log.active
        ws_log.title = "発信ログ"
        ws_log.append(["名前", "電話番号", "発信日時", "発信結果", "エラー内容", "ContactId"])

        # ======================================
        # Amazon Connectで発信実行
        # ======================================
        try:
            response = connect_client.start_outbound_voice_contact(
                InstanceId=INSTANCE_ID,
                ContactFlowId=contact_flow_id,
                SourcePhoneNumber='+14843019679',
                DestinationPhoneNumber=phone_number,
                Attributes=custom_attributes
            )
            contact_id = response['ContactId']
            logger.info(f"{phone_number} に発信成功: {contact_id}")
            ws_log.append([name, phone_number, now_str, "成功", "", contact_id])
        except Exception as e:
            # 発信失敗時のログ記録
            logger.error(f"{phone_number} への発信失敗: {str(e)}")
            ws_log.append([name, phone_number, now_str, "失敗", str(e), ""])

        # ======================================
        # 発信ログをS3に保存
        # ======================================
        output_stream = BytesIO()
        wb_log.save(output_stream)
        output_stream.seek(0)

        s3_client.put_object(
            Bucket=BUCKET_NAME,
            Key=log_key,
            Body=output_stream.read(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # ======================================
        # 最新ログのキーと現在のインデックスをS3に保存（Lambda②で使用）
        # ======================================
        LOG_OUTPUT_KEY_RECORD = f"{CALL_LOG_FOLDER}latest_log_key.json"
        log_info = {"log_key": log_key, "contact_index": contact_index}
        s3_client.put_object(
            Bucket=BUCKET_NAME,
            Key=LOG_OUTPUT_KEY_RECORD,
            Body=json.dumps(log_info).encode('utf-8'),
            ContentType='application/json'
        )

        logger.info(f"発信ログを保存: s3://{BUCKET_NAME}/{log_key}")
        logger.info(f"最新ログ情報を記録: s3://{BUCKET_NAME}/{LOG_OUTPUT_KEY_RECORD}")

        return {"statusCode": 200, "body": f"{phone_number} に発信完了。ログキー: {log_key}"}

    except Exception as e:
        # 想定外エラーのハンドリング
        logger.error(f"処理中にエラー発生: {str(e)}")
        return {"statusCode": 500, "body": f"エラー: {str(e)}"}
