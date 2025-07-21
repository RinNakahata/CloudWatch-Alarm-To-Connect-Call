# ======================================
# 標準ライブラリ
# ======================================
import json
import logging
from io import BytesIO

# ======================================
# サードパーティ製ライブラリ
# ======================================
import boto3
from openpyxl import load_workbook

# ======================================
# ログ出力設定
# ======================================
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# ======================================
# AWSサービスのクライアント生成
# ======================================
s3 = boto3.client('s3')                # S3へアクセスするためのクライアント
lambda_client = boto3.client('lambda') # 別のLambda関数を呼び出すためのクライアント

# ======================================
# 定数設定
# ======================================
BUCKET_NAME = 'amazon-connect-list'                      # 使用するS3バケット名
LOG_OUTPUT_KEY_RECORD = 'connect-call-log/latest_log_key.json'  # 最新ログ情報のS3キー


# ======================================
# Lambda関数のメイン処理
# ======================================
def lambda_handler(event, context):
    try:
        logger.info(f"受信イベント: {json.dumps(event)}")

        # CloudWatch Logs のデータをデコードしてログを取得
        message = event['awslogs']['data']
        import base64
        import gzip
        decoded = base64.b64decode(message)
        log_data = json.loads(gzip.decompress(decoded).decode('utf-8'))
        log_events = log_data['logEvents']

        # ContactId を含むログのみ抽出
        logs = [json.loads(e['message']) for e in log_events if 'ContactId' in e['message']]
        contact_id = logs[0].get("ContactId", "") if logs else ""
        logger.info(f"対象のContactId: {contact_id}")

        if not contact_id:
            raise ValueError("ログ内にContactIdが見つかりません")

        # ======================================
        # 最新のログファイル情報をS3から取得
        # ======================================
        log_info_obj = s3.get_object(Bucket=BUCKET_NAME, Key=LOG_OUTPUT_KEY_RECORD)
        log_info = json.loads(log_info_obj['Body'].read().decode('utf-8'))
        log_key = log_info["log_key"]                    # 対象のExcelファイルのS3キー
        contact_index = log_info.get("contact_index", 0) # 現在のインデックス

        # ======================================
        # Excelファイルの読み込み
        # ======================================
        s3_response = s3.get_object(Bucket=BUCKET_NAME, Key=log_key)
        wb = load_workbook(BytesIO(s3_response['Body'].read()))
        ws = wb.active  # 最初のシートを使用

        # ヘッダー行の項目と列インデックスを取得
        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

        # 「応答状況」列がなければ新たに追加
        if "応答状況" not in headers:
            ws.cell(row=1, column=ws.max_column + 1, value="応答状況")
            headers["応答状況"] = ws.max_column - 1  # 0始まりのため調整

        # ======================================
        # ContactId に一致する行を探す
        # ======================================
        matched_row = None
        for row in ws.iter_rows(min_row=2):
            if str(row[headers["ContactId"]].value) == contact_id:
                matched_row = row
                break

        if not matched_row:
            raise ValueError("一致するContactIdの行が見つかりません")

        # ======================================
        # ユーザーの入力結果（電話操作）をログから抽出
        # ======================================
        input_result = None
        for log in logs:
            if log.get("ContactFlowModuleType") == "GetUserInput":
                input_result = log.get("Results", "")
                break

        logger.info(f"ユーザー入力: {input_result}")

        # ======================================
        # 入力結果に応じて応答状況を記録し、必要なら次の人に発信
        # ======================================
        if input_result == "1":
            matched_row[headers["応答状況"]].value = "応答あり（対応可能）"
        elif input_result == "2":
            matched_row[headers["応答状況"]].value = "応答あり（対応不可）"
            logger.info("応答不可として次の人に発信")
            call_next(contact_index + 1)
        else:
            matched_row[headers["応答状況"]].value = "応答なし"
            logger.info("応答なしとして次の人に発信")
            call_next(contact_index + 1)

        # ======================================
        # 編集済みExcelファイルをS3に保存（別名で）
        # ======================================
        output_stream = BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        output_key = log_key.replace(".xlsx", "_checked.xlsx")
        s3.put_object(
            Bucket=BUCKET_NAME,
            Key=output_key,
            Body=output_stream.read(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        logger.info(f"ログ保存済み: s3://{BUCKET_NAME}/{output_key}")
        return {
            "statusCode": 200,
            "body": f"応答記録完了: {output_key}"
        }

    except Exception as e:
        logger.error(f"処理失敗: {str(e)}")
        return {
            "statusCode": 500,
            "body": f"エラー: {str(e)}"
        }


# ======================================
# 次の対象者にLambda①を使って発信を依頼
# ======================================
def call_next(next_index):
    response = lambda_client.invoke(
        FunctionName='Test_AmazonConnect',         # 呼び出すLambda関数名（発信用）
        InvocationType='Event',                    # 非同期で呼び出し
        Payload=json.dumps({"next_index": next_index}).encode('utf-8')
    )
    logger.info(f"Lambda① 呼び出し結果: {response}")
